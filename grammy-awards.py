import re
import requests
from bs4 import BeautifulSoup
import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO

# 首届格莱美于1959年举办，褒奖1958年的音乐成就
# 在官网上首届格莱美称为1958年格莱美奖
magic_number = 1957
domain = "https://www.grammy.com/"
target = f"{domain}awards/67th-annual-grammy-awards-2024"

default_avatar = "https://naras.a.bigcontent.io/v1/static/artist_default_200x200"

filename = "grammy-awards-of-all-years"

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36"
}


def main():
    json_data = fetch_and_parse(target)

    awards_data_of_all_years = []

    last_awards_data = parse_data(json_data)
    awards_data_of_all_years.extend(last_awards_data)

    awards_years_results_links = [
        domain + item["slug"]
        for item in json_data["props"]["pageProps"]["pageContent"][
            "getAwardsYearsList"
        ]["hits"]
    ]

    awards_years_results_links.remove(target)
    for link in awards_years_results_links:
        json_data = fetch_and_parse(link)
        awards_data = parse_data(json_data)
        awards_data_of_all_years.extend(awards_data)

    save_to_excel(awards_data_of_all_years, filename, False)


def save_to_excel(
    awards_data_of_all_years: list, path: str, with_img: bool = False
) -> None:
    df = pd.DataFrame(awards_data_of_all_years)
    df.to_excel(f"{path}.xlsx", sheet_name="grammy-awards", index=False)

    wb = load_workbook(f"{path}.xlsx")
    ws = wb["grammy-awards"]

    if with_img:
        ws.cell(row=1, column=len(df.columns) + 1, value="image")
        ws.column_dimensions[chr(65 + len(df.columns) + 1)].width = 50
        for i in range(1, len(awards_data_of_all_years) + 2):
            ws.row_dimensions[i].height = 50

        img_cache = {}

        for idx, row in df.iterrows():
            image_url = row["avatar"]
            if image_url:
                if image_url in img_cache.keys():
                    cached_img = Image(BytesIO(img_cache[image_url]))
                    cached_img.width, cached_img.height = 40, 40
                    ws.add_image(cached_img, f"{chr(65 + len(df.columns))}{idx + 2}")
                    print(
                        f"\033[32m{idx + 1}/{len(awards_data_of_all_years)} Image {image_url} inserted from cache\033[0m"
                    )
                else:
                    try:
                        response = requests.get(image_url, headers=headers)
                        response.raise_for_status()
                        img = Image(BytesIO(response.content))
                        img.width, img.height = 40, 40
                        ws.add_image(img, f"{chr(65 + len(df.columns))}{idx + 2}")
                        img_cache[image_url] = response.content
                        print(
                            f"{idx + 1}/{len(awards_data_of_all_years)} Image {image_url} inserted"
                        )
                    except Exception as e:
                        default_img = Image(BytesIO(img_cache[default_avatar]))
                        default_img.width, default_img.height = 40, 40
                        ws.add_image(
                            default_img, f"{chr(65 + len(df.columns))}{idx + 2}"
                        )
                        print(
                            f"\033[31mFailed to insert image {image_url} due to {e}; Use default avatar instead\033[31m"
                        )
    wb.save(f"{path}.xlsx")
    print(f"Data saved to {os.getcwd()}\\{path}.xlsx")


def fetch_and_parse(url: str) -> dict:
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print("Failed to retrieve data")
        return

    soup = BeautifulSoup(response.content, "html.parser")
    script_tag = soup.find("script", id="__NEXT_DATA__")

    if script_tag is None:
        print("Failed to find tag <script>")
        return

    json_data = json.loads(script_tag.string)
    if json_data is None:
        print("Failed to retrieve JSON data")

    print(f"Successfully fetched json data from {url}")

    return json_data


def parse_data(json_data: dict) -> list:
    awards_years = json_data["props"]["pageProps"]["pageContent"]["getAwardsYears"][
        "hits"
    ][0]

    edition = get_th_order(awards_years["title"])  # 67 "67th Annual GRAMMY Awards"
    holdingYear = magic_number + edition  # 2025 举办年份

    print(f"\033[34mWelcome to the {edition}th Annual GRAMMY Awards!\033[0m")

    awards_list = []

    for award_detail in awards_years["categoryDetails"]:
        award_name = award_detail["title"][0]["name"]  # "Record Of The Year"
        nominations = award_detail["nominations"]

        single_award_list = []

        winner = ""

        for nomination in nominations:
            part1 = strip_only_once(
                re.sub(r"\\*", "", nomination["displayLine1"]), '"'
            )  # "Not Like Us"
            isWinner: bool = nomination["isWinner"]  # true
            nomineeOrder = nomination["nomineeOrder"]  # 6

            part2 = nomination[
                "displayLine2"
            ]  # "\u003ca href=\"/artists/kendrick-lamar/17949\"\u003eKendrick Lamar\u003c/a\u003e"
            if award_name == "Best New Artist":
                part2 = nomination["title"]
            if (part2 is None or part2 == "") and nomination[
                "displayLine3"
            ] is not None:
                # Kayleigh Rose Amstutz, Daniel Nigro \u0026 Justin Tranter, songwriters (Chappell Roan)
                match = re.search(r"\((.*?)\)", nomination["displayLine3"])
                if match:
                    part2 = match.group(1)

            part3 = nomination[
                "displayLine3"
            ]  # "Mustard, Sean Momberger \u0026 Sounwave, producers; Ray Charles Brown Jr. \u0026 Johnathan Turner, engineers/mixers; Nicolas de Porcel, mastering engineer"

            tivoInfo = nomination["creditedArtists"][0]["tivoInfo"]
            if tivoInfo is not None:
                avatar = tivoInfo["damDynamic"]
            if avatar is None or avatar == "":
                avatar = default_avatar

            single_award_list.append(
                {
                    "year": holdingYear,
                    "edition": edition,
                    "award": award_name,
                    "part1": part1.replace("\r", "").replace("\n", ""),
                    "part2": remove_tags(part2).replace("(", "").replace(")", ""),
                    "part3": replace_unicode(part3),
                    "win": isWinner,
                    "order": nomineeOrder,
                    "avatar": avatar,
                }
            )

            if isWinner:
                winner = part1
                if part2 is not None and part2 != "":
                    winner = f"{part1} - {remove_tags(part2)}"

        print(f"{edition}th {award_name}: {winner}")

        single_award_list.sort(key=lambda x: x["order"])
        awards_list.extend(single_award_list)

    return awards_list


def get_th_order(s: str) -> int:
    match = re.search(r"\d+", s)
    if match:
        number = int(match.group())
        return number
    else:
        print("No number found in the string.")


def remove_tags(s: str) -> str:
    if s is None or s == "":
        return ""
    html_string = s.replace("\u003c", "<").replace("\u003e", ">")
    cleaned_string = re.sub(r"<.*?>", "", html_string)
    return cleaned_string


def replace_unicode(s) -> str:
    if s is None or s == "":
        return ""
    return re.sub(r"\\u([0-9a-fA-F]{4})", lambda match: chr(int(match.group(1), 16)), s)


def strip_only_once(s: str, ch: chr) -> str:
    if s.startswith(ch) and s.endswith(ch):
        return s[1:-1]
    else:
        return s


if __name__ == "__main__":
    main()
